from openpyxl import Workbook, load_workbook
from string import ascii_uppercase
from openpyxl.styles import PatternFill, Font
from os.path import exists
from sys import exit


def get_input():
    # make sure all neccessary files for the program exist
    missing_file = False
    for i in ["input.txt", "departments.txt", "class values.xlsx"]:
        if not exists(i):
            print(f"error: no file named {i} found")
            missing_file = True
    if missing_file:
        print("program terminated, no changes have been made")
        input("press enter to exit")
        exit()

    # get and interpret input.txt
    file = open("input.txt", "r")
    txt = file.read()
    strings_lists = txt.split("\n")
    del strings_lists[0]
    new_list = []
    for i in range(len(strings_lists)):
        l = strings_lists[i].split(",")
        l = [j.strip('"') for j in l]
        if len(l) == 4:
            new_list.append(l)
    return new_list


def get_xl_col(n):
    # get the excel column given numeric values
    if n > 25:
        letter = ascii_uppercase[n // 26 - 1] + ascii_uppercase[n % 26]
    else:
        letter = ascii_uppercase[n]
    return letter


def initialize_sheet():
    # set width
    for i in range(10 * w + 1):
        sheet.column_dimensions[get_xl_col(i)].width = 7.33

    # write compartments
    for i in range(10):
        sheet[get_xl_col(i * w + 1) + '1'] = ["ELA","Fine Arts","Math","Phys. Ed","Science","Social Science","PAA","Frech Imm. & Second Languages","EAL & Support","Totals"][i]

    # write periods
    j = 2
    rows = []
    for i in range(12):
        sheet['A' + str(j)] = ['P1', 'P2', 'P3', 'Lunch', 'P4', 'P5', 'P1', 'P2', 'P3', 'Lunch', 'P4', 'P5'][i]
        rows.append(j)
        if i == 3 or i == 9:
            j += lunch_space
        else:
            j += h

    # fill colours
    for i in range(1, 10 * w + 1):
        if (i - 1) % w <= 3:
            for j in rows:
                for k in range(j, j + h - 1):
                    if (rows.index(j) == 3 or rows.index(j) == 9) and k > j:
                        break
                    hex_color = ["FFCCCB", "FFFFE0", "ADD8E6", "90EE90"][(i - 1) % w]
                    sheet[get_xl_col(i) + str(k)].fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # grey boarders between departments
    for i in range(5, 9 * w + 5, w):
        sheet.column_dimensions[get_xl_col(i)].width = 2
        for j in sheet[get_xl_col(i)]:
            j.fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")

    # boarders between each period
    if w >= 5:
        for i in range(1, 12):
            sheet.row_dimensions[rows[i] - 1].height = 10
            for j in range(9 * w + 5):
                sheet[get_xl_col(j) + str(rows[i] - 1)].fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
        sheet.row_dimensions[5 * h + lunch_space + 1].height = 20
        for i in range(9 * w + 5):
            sheet[get_xl_col(i) + str(5 * h + lunch_space + 1)].fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
    
    # color indication
    for i in range(1, 5):
        color = ["FFCCCB", "FFFFE0", "ADD8E6", "90EE90"][i - 1]
        sheet[get_xl_col(i) + str(10 * h + lunch_space * 2 + 2)].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet[get_xl_col(i) + str(10 * h + lunch_space * 2 + 3)] = ["Gr. 9", "Gr. 10", "Gr. 11", "Gr. 12"][i - 1]
    return rows


def register(i, compartment, index):
    # registering a class to its cell
    j = 0
    while not sheet[get_xl_col(compartment) + str(index + j)].value is None:
        j += 1

    # purple color for grouped classes
    if "grp" in i[0]:
        sheet[get_xl_col(compartment) + str(index + j)] = i[0]
        sheet[get_xl_col(compartment) + str(index + j)].font = Font(color="A020F0")
    else:
        sheet[get_xl_col(compartment) + str(index + j)] = i[0].rsplit("-", 1)[0]

    # red color for full year classes
    if i[1][1] == "Y":
        sheet[get_xl_col(compartment) + str(index + j)].font = Font(color="FF0000")


def totals_register(i, class_list, totals, period, grade_val, val_sheet):
    # adding to totals list according to class values.xlsx
    if i[3] == "ADMIN" or i[3] == "EAL":
        return
    the_class = i[0].rsplit("-", 1)[0]
    if not the_class in class_list:
        totals[period][grade_val] += 1 - ("grp" in i[0]) * 0.5
    else:
        string = val_sheet["F" + str(class_list.index(the_class) + 2)].value
        if string is None or i[1][1] in str(string):
            for i in range(1, 5):
                value = val_sheet[get_xl_col(i) + str(class_list.index(the_class) + 2)].value
                if not value is None:
                    totals[period][i - 1] += float(value)


def fill_sheet():
    # removing lunch values
    del rows[3]
    del rows[8]

    # initializing totals list
    totals = []
    for _ in range(12):
        totals.append([0, 0, 0, 0])

    # interpreting and storing the contents of departments.txt
    content = open("departments.txt", "r").read()
    mp = eval(content)
    for i in mp:
        if not type(i) is str or len(i) != 3 or not i.isalpha() or not type(mp[i]) is int or mp[i] > 8 or mp[i] < 0:
            print("contents of departments.txt not in the correct format, program terminated, no changes have been made")
            input("press enter to exit")
            exit()

    # loading and storing class values.xlsx
    val_sheet = load_workbook(filename="class values.xlsx")["Sheet1"]

    # identify the classes in class values.xlsx
    class_list = []
    i = 2
    while not val_sheet["A" + str(i)].value is None:
        class_list.append(val_sheet["A" + str(i)].value)
        i += 1
    # putting classes in their cells
    for i in strings_lists:
        # determining the horizontal location
        if i[0][5] == "F":
            grade_val = int(i[0][3:5]) // 10
            compartment = 7 * w + 1 + grade_val
        elif i[3] == "ADMIN":
            grade_val = 0
            compartment = 8 * w + 1
        else:
            grade_val = int(i[0][3:5]) // 10
            if i[0][:3] in mp:
                compartment = mp[i[0][:3]] * w + 1 + grade_val
            else:
                print("there was an error finding a department for this class", i[0])
                continue

        # determining the vertical location
        period = i[2][0]
        semester = i[1][1]
        if period == "L":
            index = rows[3] - lunch_space
        else:
            index = rows[int(period) - 1]

        # writing the class into its cell
        if semester == "1" or semester == "Y":
            register(i, compartment, index)
            if period != "L":
                totals_register(i, class_list, totals, int(period) - 1, grade_val, val_sheet)
            else:
                totals_register(i, class_list, totals, 10, grade_val, val_sheet)
        if semester == "2" or semester == "Y":
            register(i, compartment, index + 5 * h + lunch_space)
            if period != "L":
                totals_register(i, class_list, totals, int(period) + 4, grade_val, val_sheet)
            else:
                totals_register(i, class_list, totals, 11, grade_val, val_sheet)

    # totals registration
    for i in range(10):
        for j in range(4):
            sheet[get_xl_col(9 * w + 1 + j) + str(rows[i])] = totals[i][j]
    for i in range(4):
        sheet[get_xl_col(9 * w + 1 + i) + str(rows[3] - lunch_space)] = totals[10][i]
        sheet[get_xl_col(9 * w + 1 + i) + str(rows[8] - lunch_space)] = totals[11][i]


strings_lists = get_input()
schedule = Workbook()
sheet = schedule.active
w = 5
h = 6
lunch_space = 2
rows = initialize_sheet()
fill_sheet()
schedule.save("output.xlsx")
input("program has run successfully, press enter to exit")
